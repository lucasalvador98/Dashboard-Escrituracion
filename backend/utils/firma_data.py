import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

FIRMA_COLUMNS = [
    ("Seccional", 10),
    ("BARRIO", 27),
    ("Manzana plano", 14),
    ("Lote plano", 11),
    ("Manzana oficial", 15),
    ("Lote oficial", 12),
    ("BENEFICIARIO", 42),
    ("Dni", 15),
    ("Telefono", 25),
    ("COTITULAR - Nombre y Apellido", 35),
    ("COTITULAR - DNI", 15),
    ("Telefono", 18),
]

THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)

HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def _get(item, *keys):
    for k in keys:
        v = item.get(k)
        if v is not None and v != "":
            return v
    return ""


def generar_firma_excel(datos, titulo="", fecha="", hora="", lugar="", escribano_nombre="", escribano_tel="", escribano_mail=""):
    """
    Genera un Excel con formato FIRMA (En Trámite).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    col_count = len(FIRMA_COLUMNS)

    for i, (_, width) in enumerate(FIRMA_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Fila 1: Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    cell = ws.cell(1, 1, titulo or "")
    cell.font = Font(name="Arial", size=18, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # Fila 2: Fecha / Hora / Lugar
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    ws.cell(2, 1, f"FECHA DE FIRMA: {fecha}").font = Font(name="Arial", size=11, bold=True)

    ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=7)
    ws.cell(2, 5, f"HORA: {hora}HS").font = Font(name="Arial", size=11, bold=True)

    ws.merge_cells(start_row=2, start_column=8, end_row=2, end_column=col_count)
    ws.cell(2, 8, f"LUGAR: {lugar}").font = Font(name="Arial", size=11, bold=True)

    # Fila 3: Localidad/Barrio + Escribano
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)
    ws.cell(3, 1, "LOCALIDAD / BARRIO").font = Font(name="Arial", size=11, bold=True)

    ws.merge_cells(start_row=3, start_column=5, end_row=3, end_column=col_count)
    ws.cell(3, 5, "ESCRIBANO").font = Font(name="Arial", size=11, bold=True)

    # Fila 4: Contacto
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=4)
    ws.cell(4, 1, "CONTACTO:").font = Font(name="Arial", size=11, bold=True)

    ws.merge_cells(start_row=4, start_column=5, end_row=4, end_column=col_count)
    ws.cell(4, 5, f"ESCRIBANO: {escribano_nombre}").font = Font(name="Arial", size=11)

    # Fila 5: Teléfono
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=4)
    ws.cell(5, 1, "TELEFONO:").font = Font(name="Arial", size=11, bold=True)

    ws.merge_cells(start_row=5, start_column=5, end_row=5, end_column=col_count)
    ws.cell(5, 5, f"TELEFONO: {escribano_tel}").font = Font(name="Arial", size=11)

    # Fila 6: Mail
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=4)
    ws.cell(6, 1, "MAIL:").font = Font(name="Arial", size=11, bold=True)

    ws.merge_cells(start_row=6, start_column=5, end_row=6, end_column=col_count)
    ws.cell(6, 5, f"MAIL: {escribano_mail}").font = Font(name="Arial", size=11)

    # Fila 7: Headers
    for i, (label, _) in enumerate(FIRMA_COLUMNS, 1):
        cell = ws.cell(7, i, label)
        cell.font = Font(name="Arial", size=11, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        cell.fill = HEADER_FILL

    # Filas de datos
    for idx, item in enumerate(datos, 1):
        row_num = idx + 7

        ws.cell(row_num, 1, _get(item, "Seccional", "seccional")).font = Font(name="Arial", size=12)
        ws.cell(row_num, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_num, 1).border = THIN_BORDER

        ws.cell(row_num, 2, _get(item, "Barrio")).font = Font(name="Arial", size=12)
        ws.cell(row_num, 2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_num, 2).border = THIN_BORDER

        ws.cell(row_num, 3, _get(item, "Mza. Plano", "Mza. Oficial", "Mza")).font = Font(name="Arial", size=12)
        ws.cell(row_num, 3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_num, 3).border = THIN_BORDER

        ws.cell(row_num, 4, _get(item, "Lote Plano", "Lote oficial", "Lote")).font = Font(name="Arial", size=12)
        ws.cell(row_num, 4).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_num, 4).border = THIN_BORDER

        ws.cell(row_num, 5, _get(item, "Mza. Oficial", "Mza. Plano", "Mza")).font = Font(name="Arial", size=12)
        ws.cell(row_num, 5).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_num, 5).border = THIN_BORDER

        ws.cell(row_num, 6, _get(item, "Lote oficial", "Lote Plano", "Lote")).font = Font(name="Arial", size=12)
        ws.cell(row_num, 6).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_num, 6).border = THIN_BORDER

        ws.cell(row_num, 7, _get(item, "Beneficiarios", "Beneficiario", "APELLIDO Y NOMBRE")).font = Font(name="Arial", size=12)
        ws.cell(row_num, 7).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_num, 7).border = THIN_BORDER

        ws.cell(row_num, 8, _get(item, "DNI", "dni")).font = Font(name="Arial", size=12)
        ws.cell(row_num, 8).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_num, 8).border = THIN_BORDER

        ws.cell(row_num, 9, _get(item, "Telefono", "telefono", "Teléfono")).font = Font(name="Arial", size=12)
        ws.cell(row_num, 9).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_num, 9).border = THIN_BORDER

        ws.cell(row_num, 10, _get(item, "COTITULAR Nombre y Apellido", "COTITULAR - Nombre y Apellido", "Cotitular")).font = Font(name="Arial", size=12)
        ws.cell(row_num, 10).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_num, 10).border = THIN_BORDER

        ws.cell(row_num, 11, _get(item, "COTITULAR DNI", "COTITULAR - DNI", "CotitularDNI")).font = Font(name="Arial", size=12)
        ws.cell(row_num, 11).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_num, 11).border = THIN_BORDER

        ws.cell(row_num, 12, _get(item, "COTITULAR Telefono", "Tel. Cotitular", "TelefonoCotitular")).font = Font(name="Arial", size=12)
        ws.cell(row_num, 12).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_num, 12).border = THIN_BORDER

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
