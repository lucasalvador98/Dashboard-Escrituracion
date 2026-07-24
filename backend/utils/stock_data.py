import io
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("NRO", 6),
    ("BARRIO", 27),
    ("MZA", 7),
    ("LOTE", 7),
    ("APELLIDO Y NOMBRE", 42),
    ("DNI", 20),
    ("Teléfono", 21),
    ("COTITULAR - Nombre y Apellido", 35),
    ("COTITULAR - DNI", 24),
    ("Tel. Cotitular", 41),
    ("ASISTENCIA", 15),
]

THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)

HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def _get(item, *keys):
    """Busca múltiples variantes del nombre de campo."""
    for k in keys:
        v = item.get(k)
        if v is not None and v != "":
            return v
    return ""


def generar_excel(datos, titulo="", subtitulo=""):
    """
    Genera un Excel en memoria con el formato del modelo de VILLA CARLOS PAZ.
    - datos: lista de dicts con keys: Barrio, Beneficiario, DNI, etc.
    - titulo: texto del título (ej: nombre del departamento/localidad)
    - subtitulo: texto del subtítulo
    Devuelve un BytesIO con el .xlsx listo para descargar.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    col_count = len(COLUMNS)

    # ── Anchos de columna ──
    for i, (_, width) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ── Fila 1: Título (merged) ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    cell_titulo = ws.cell(1, 1, titulo or "")
    cell_titulo.font = Font(name="Arial", size=18, bold=True)
    cell_titulo.alignment = Alignment(horizontal="center", vertical="center")

    # ── Fila 2: Subtítulo (merged) ──
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    cell_sub = ws.cell(2, 1, subtitulo or "")
    cell_sub.font = Font(name="Arial", size=14, bold=True)
    cell_sub.alignment = Alignment(horizontal="center", vertical="center")

    # ── Fila 3: Headers ──
    for i, (label, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(3, i, label)
        cell.font = Font(name="Arial", size=11, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        cell.fill = HEADER_FILL

    # ── Filas de datos ──
    for idx, item in enumerate(datos, 1):
        row_num = idx + 3

        ws.cell(row_num, 1, idx).font = Font(name="Arial", size=12)
        ws.cell(row_num, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_num, 1).border = THIN_BORDER

        ws.cell(row_num, 2, item.get("Barrio") or "").font = Font(name="Arial", size=12)
        ws.cell(row_num, 2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_num, 2).border = THIN_BORDER

        ws.cell(row_num, 3, _get(item, "Mza. Plano", "Mza. Oficial", "Mza", "MZA", "mza")).font = Font(name="Arial", size=12)
        ws.cell(row_num, 3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_num, 3).border = THIN_BORDER

        ws.cell(row_num, 4, _get(item, "Lote Plano", "Lote oficial", "Lote Oficial", "Lote", "LOTE")).font = Font(name="Arial", size=12)
        ws.cell(row_num, 4).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_num, 4).border = THIN_BORDER

        ws.cell(row_num, 5, _get(item, "Beneficiarios", "Beneficiario", "APELLIDO Y NOMBRE", "ApellidoYNombre", "Nombre")).font = Font(name="Arial", size=12)
        ws.cell(row_num, 5).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_num, 5).border = THIN_BORDER

        ws.cell(row_num, 6, _get(item, "DNI", "dni", "documento")).font = Font(name="Arial", size=12)
        ws.cell(row_num, 6).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_num, 6).border = THIN_BORDER

        ws.cell(row_num, 7, _get(item, "Telefono", "telefono", "Teléfono")).font = Font(name="Arial", size=12)
        ws.cell(row_num, 7).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_num, 7).border = THIN_BORDER

        ws.cell(row_num, 8, _get(item, "COTITULAR Nombre y Apellido", "COTITULAR - Nombre y Apellido", "Cotitular")).font = Font(name="Arial", size=12)
        ws.cell(row_num, 8).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_num, 8).border = THIN_BORDER

        ws.cell(row_num, 9, _get(item, "COTITULAR DNI", "COTITULAR - DNI", "CotitularDNI")).font = Font(name="Arial", size=12)
        ws.cell(row_num, 9).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_num, 9).border = THIN_BORDER

        ws.cell(row_num, 10, _get(item, "COTITULAR Telefono", "Tel. Cotitular", "TelefonoCotitular")).font = Font(name="Arial", size=12)
        ws.cell(row_num, 10).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_num, 10).border = THIN_BORDER

        ws.cell(row_num, 11, _get(item, "Asistencia", "ASISTENCIA")).font = Font(name="Arial", size=12)
        ws.cell(row_num, 11).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_num, 11).border = THIN_BORDER

    # Guardar en memoria
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
